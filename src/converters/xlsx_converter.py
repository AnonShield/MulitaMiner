"""XLSX (Excel) converter."""

import os
from typing import List, Dict, Any, Optional
from datetime import datetime

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from .base_converter import BaseConverter


class XLSXConverter(BaseConverter):
    """Converts extraction JSON to a styled XLSX workbook."""

    def __init__(self):
        super().__init__()

        if not PANDAS_AVAILABLE:
            raise ImportError("pandas is required for XLSX conversion. Install with: pip install pandas")
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for XLSX conversion. Install with: pip install openpyxl")

    def get_format_name(self) -> str:
        return "XLSX"

    def create_styled_workbook(self, data: List[Dict[str, Any]]) -> Workbook:
        """Build the styled workbook (data sheet + metadata sheet)."""
        # Lists become multi-line cells, None becomes empty
        def normalize_for_xlsx(value):
            if value is None:
                return ''
            if isinstance(value, list):
                return '\n'.join(str(v) for v in value)
            return str(value)

        normalized_data = []
        for item in data:
            normalized_item = {}
            for key, value in item.items():
                normalized_item[key] = normalize_for_xlsx(value)
            normalized_data.append(normalized_item)
        df = pd.DataFrame(normalized_data)

        # Known fields first, remaining columns keep their order
        column_order = []
        for field in self.supported_fields:
            if field in df.columns:
                column_order.append(field)

        for col in df.columns:
            if col not in column_order:
                column_order.append(col)

        df = df[column_order]

        wb = Workbook()
        ws = wb.active
        ws.title = "Vulnerabilities"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass

            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        metadata_ws = wb.create_sheet("Metadata")
        metadata_ws['A1'] = "Report Generation Info"
        metadata_ws['A1'].font = Font(bold=True, size=14)

        metadata_ws['A3'] = "Generated on:"
        metadata_ws['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        metadata_ws['A4'] = "Total vulnerabilities:"
        metadata_ws['B4'] = len(data)

        metadata_ws['A5'] = "Converter:"
        metadata_ws['B5'] = f"{self.get_format_name()} Converter"

        if data:
            severity_counts = {}
            for item in data:
                severity = item.get('Risk', item.get('severity', 'Unknown'))
                severity_counts[severity] = severity_counts.get(severity, 0) + 1

            metadata_ws['A7'] = "Severity Distribution:"
            metadata_ws['A7'].font = Font(bold=True)

            row = 8
            for severity, count in severity_counts.items():
                metadata_ws[f'A{row}'] = f"{severity}:"
                metadata_ws[f'B{row}'] = count
                row += 1

        return wb

    def convert(self, json_file_path: str, output_file_path: Optional[str] = None) -> str:
        """Convert a JSON file to XLSX and return the output path."""
        if output_file_path is None:
            output_file_path = self.get_output_filename(json_file_path, "xlsx")

        # Reuse the existing XLSX when it is newer than the JSON (cache)
        if os.path.exists(output_file_path):
            json_mtime = os.path.getmtime(json_file_path)
            xlsx_mtime = os.path.getmtime(output_file_path)

            if xlsx_mtime >= json_mtime:
                print(f"Using existing XLSX file (cache): {output_file_path}")
                return output_file_path
            else:
                print(f"JSON file is newer than XLSX, reconverting: {json_file_path}")

        data = self.load_json_data(json_file_path)

        if not self.validate_data(data):
            raise ValueError("Invalid JSON data")

        if not data:
            print("Warning: No vulnerabilities found in JSON")
            data = [{"name": "No vulnerabilities found", "description": "Empty report"}]

        try:
            wb = self.create_styled_workbook(data)
            wb.save(output_file_path)
            print(f"XLSX file created successfully: {output_file_path}")
            print(f"Total vulnerabilities: {len(data)}")
            return output_file_path
        except Exception as e:
            raise Exception(f"Error creating XLSX file: {e}")


def convert_json_to_xlsx(json_file_path: str, output_file_path: Optional[str] = None) -> str:
    """Direct JSON to XLSX conversion helper."""
    converter = XLSXConverter()
    return converter.convert(json_file_path, output_file_path)
